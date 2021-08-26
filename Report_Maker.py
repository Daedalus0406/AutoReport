import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis


# start_date = date[0]
# rm.report_maker(report_3800, report_3600, report_2400, dates[0])

def weekly_report_maker(report_38, report_36, report_24, report_date, year, week_num):
    # excel產生
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'VPI3800'
    ws2 = wb.create_sheet('VPI3600')
    ws3 = wb.create_sheet('VPI2400')

    report_38.index.name = "日期"
    for r in dataframe_to_rows(report_38.reset_index(), index=False):
        ws1.append(r)

    report_36.index.name = "日期"
    for r in dataframe_to_rows(report_36.reset_index(), index=False):
        ws2.append(r)

    report_24.index.name = "日期"
    for r in dataframe_to_rows(report_24.reset_index(), index=False):
        ws3.append(r)

    def charts(ws, ws_title):
        ws.insert_rows(1)
        ws.merge_cells('A1:I1')
        ws['A1'] = ws_title + ' 設備稼動週報表'
        ws['A1'].font = Font(size=15, bold=True)
        ws['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # center all cells
        for col in ws.columns:
            for cell in col:
                # openpyxl styles aren't mutable,
                # so you have to create a copy of the style, modify the copy, then set it back
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj

        row = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        for i in row:
            ws.row_dimensions[i].height = 20

        col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for j in col:
            ws.column_dimensions[j].width = 15

        # 圖表
        # Chart with date axis
        dates = Reference(ws, min_col=1, min_row=3, max_row=9)
        # 折線
        c1 = LineChart()
        c1.title = "設備工時"
        c1.style = 12
        c1.y_axis.title = "分"
        c1.y_axis.crossAx = 500
        c1.x_axis = DateAxis(crossAx=100)
        c1.x_axis.number_format = 'd-mmm'
        c1.x_axis.majorTimeUnit = "days"
        c1.x_axis.title = "日期"

        data1 = Reference(ws, min_col=5, max_col=8, min_row=2, max_row=9)
        c1.add_data(data1, titles_from_data=True)
        c1.set_categories(dates)
        ws.add_chart(c1, "A12")

        # 直條
        c2 = BarChart()
        c2.type = "col"
        c2.style = 10
        c2.title = "設備效能"
        c2.y_axis.title = '%'
        c2.x_axis.title = '日期'

        data2 = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=9)
        c2.add_data(data2, titles_from_data=True)
        c2.set_categories(dates)
        c2.shape = 4
        ws.add_chart(c2, "H12")

    charts(ws1, ws1.title)
    charts(ws2, ws2.title)
    charts(ws3, ws3.title)
    wb.save("VPI含浸爐_設備稼動週報表_" + year + "_" + week_num + "_" + report_date + ".xlsx")
    print("報表生成完畢")


def monthly_report_maker(report_38, report_36, report_24, report_date, year, month, monthrange):
    monthrange = monthrange+2
    # excel產生
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'VPI3800'
    ws2 = wb.create_sheet('VPI3600')
    ws3 = wb.create_sheet('VPI2400')

    report_38.index.name = "日期"
    for r in dataframe_to_rows(report_38.reset_index(), index=False):
        ws1.append(r)

    report_36.index.name = "日期"
    for r in dataframe_to_rows(report_36.reset_index(), index=False):
        ws2.append(r)

    report_24.index.name = "日期"
    for r in dataframe_to_rows(report_24.reset_index(), index=False):
        ws3.append(r)

    def charts(ws, ws_title):
        ws.insert_rows(1)
        ws.merge_cells('A1:I1')
        ws['A1'] = ws_title + ' 設備稼動月報表'
        ws['A1'].font = Font(size=15, bold=True)
        ws['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # center all cells
        for col in ws.columns:
            for cell in col:
                # openpyxl styles aren't mutable,
                # so you have to create a copy of the style, modify the copy, then set it back
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj

        row = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        for i in row:
            ws.row_dimensions[i].height = 20

        col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for j in col:
            ws.column_dimensions[j].width = 15

        # 圖表
        # Chart with date axis
        dates = Reference(ws, min_col=1, min_row=3, max_row=monthrange)
        # 折線
        c1 = LineChart()
        c1.title = "設備工時"
        c1.style = 12
        c1.y_axis.title = "分"
        c1.y_axis.crossAx = 500
        c1.x_axis = DateAxis(crossAx=100)
        c1.x_axis.number_format = 'd-mmm'
        c1.x_axis.majorTimeUnit = "days"
        c1.x_axis.title = "日期"

        data1 = Reference(ws, min_col=5, max_col=8, min_row=2, max_row=monthrange)
        c1.add_data(data1, titles_from_data=True)
        c1.set_categories(dates)
        c1.height = 10
        c1.width = 20
        loca = monthrange+3
        loca_c1 = "A"+str(loca)
        ws.add_chart(c1, loca_c1)

        # 直條
        c2 = BarChart()
        c2.type = "col"
        c2.style = 10
        c2.title = "設備效能"
        c2.y_axis.title = '%'
        c2.x_axis.title = '日期'

        data2 = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=monthrange)
        c2.add_data(data2, titles_from_data=True)
        c2.set_categories(dates)
        c2.height = 10
        c2.width = 20
        loca_c2 = "H"+str(loca)
        ws.add_chart(c2, loca_c2)

    charts(ws1, ws1.title)
    charts(ws2, ws2.title)
    charts(ws3, ws3.title)

    wb.save("VPI含浸爐_設備稼動月報表_" + year + "_" + month + "_" + report_date + ".xlsx")
    print("報表生成完畢")